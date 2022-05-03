from operator import countOf
import re
import json
import urllib3
#from pyzabbix import ZabbixAPI, ZabbixAPIException
from pyzabbix.api import ZabbixAPI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


zapi = ZabbixAPI("https://zabbix/api_jsonrpc.php", user="", password="")

print("Connected to Zabbix API Version %s" % zapi.api_version())

host_name_list = []
items_list = []
triggers_list = []

dest_filename = "C:\\Users\\sr.krasnov\\Documents\\new project\\triggers_audit.xlsx"
ac_name = 'Jira/Confluence'

wb = Workbook()
worksheet = wb.active
worksheet.title = "Welcome"
img = Image("C:\\Users\\sr.krasnov\\Documents\\new project\\atl.site.logo")
worksheet.add_image(img, 'A1')
worksheet['A4'] = 'Выгрузка метрик и триггеров из системы мониторинга Zabbix по AC "' + ac_name + '"'
worksheet['A4'].font = Font(b=True)
wb.save(filename = dest_filename)

host_list_xlsx = load_workbook("C:\\Users\\sr.krasnov\\Documents\\new project\\AD_servers.xlsx")
source = host_list_xlsx["1"]
for cell in source['A']:
    cell.value = re.sub(r"None", "", str(cell.value))
    if re.findall(r"\w+", str(cell.value)):
        host_name_list.append(str(cell.value))
    else:
        None

open("no_host_list.txt", "w").close()


for hostname in host_name_list:

    print ('Выгрузка по: ' + hostname)        
    cell_count = 1
    audit_file = load_workbook(filename = dest_filename)
    audit_file.create_sheet(title=hostname)
    worksheet = audit_file[hostname]
    worksheet["A"+str(cell_count)] = 'Item name'
    worksheet["A"+str(cell_count)].font = Font(b=True)
    worksheet["B"+str(cell_count)] = 'Item Key_'
    worksheet["B"+str(cell_count)].font = Font(b=True)
    worksheet["C"+str(cell_count)] = 'Item Delay'
    worksheet["C"+str(cell_count)].font = Font(b=True)
    worksheet["D"+str(cell_count)] = 'Item status'
    worksheet["D"+str(cell_count)].font = Font(b=True)
    worksheet["E"+str(cell_count)] = 'Trigger description'
    worksheet["E"+str(cell_count)].font = Font(b=True)
    worksheet["F"+str(cell_count)] = 'Trigger severity'
    worksheet["F"+str(cell_count)].font = Font(b=True)
    worksheet["G"+str(cell_count)] = 'Trigger status'
    worksheet["G"+str(cell_count)].font = Font(b=True)
    worksheet["H"+str(cell_count)] = 'Trigger expression'
    worksheet["H"+str(cell_count)].font = Font(b=True)
    worksheet["I"+str(cell_count)] = 'IF SQL Item select or other params'
    worksheet["I"+str(cell_count)].font = Font(b=True)

    audit_file.save(filename = dest_filename)

    get_hostid = zapi.host.get(filter={"name":str(hostname)}, output=["hostid"])
    print (get_hostid)
    if re.findall(r"\[\]", str(get_hostid)):
        no_host_list = open("no_host_list.txt", "a")
        no_host_list.write(str(hostname)+"\n")
        no_host_list.close()
    else:
        cell_count += 1
        audit_file = load_workbook(filename = dest_filename)
        worksheet = audit_file[hostname]
#        worksheet["A" + str(cell_count)] = str(hostname)
        audit_file.save(filename = dest_filename)
        dict_hostid = dict(enumerate(get_hostid))
        clear_hostid = (dict_hostid[0].get("hostid"))
        host_get_items = zapi.item.get(hostids=str(clear_hostid), expandComment=True, expandDescription=True, expandExpression=True, selectTags=["extend"], selectTriggers=True)
        item = re.findall(r"{'itemid'.*?'tags':.*?}", str(host_get_items))
        for item in host_get_items:
            if item:
                item_to_json = json.dumps(item)
                items_list.append(str(item))
                item_to_json = json.loads(item_to_json)
                print (str(item_to_json["name"]))
#                worksheet["B"+str(cell_count)] = str(item_to_json["itemid"])
                worksheet["A"+str(cell_count)] = str(item_to_json["name"])
                worksheet["A"+str(cell_count)].font = Font(size=9)
                worksheet["B"+str(cell_count)] = str(item_to_json["key_"])
                worksheet["B"+str(cell_count)].font = Font(size=9)
                worksheet["C"+str(cell_count)] = str(item_to_json["delay"])
                worksheet["C"+str(cell_count)].font = Font(size=9)
                worksheet["D"+str(cell_count)] = 'Enable' if str(item_to_json["status"]) == '0' else 'Disable'
                worksheet["D"+str(cell_count)].font = Font(size=9)
                if re.findall(r"db.odbc.*?(?=\[)", str(item_to_json["key_"])):
                    worksheet["I"+str(cell_count)] = 'No' if str(item_to_json["params"]) == '0' else str(item_to_json["params"])
                    worksheet["I"+str(cell_count)].font = Font(size=9)
                item_triggers_list = zapi.trigger.get(hostids=str(clear_hostid), itemids=str(item_to_json["itemid"]), expandComment=True, expandDescription=True, expandExpression=True, selectTags=["extend"])
                trigger = re.findall(r"{'triggerid'.*?'tags':.*?}", str(item_triggers_list))
                trigger_count = 0
                for trigger in item_triggers_list:
                    if trigger:
                        trigger_to_json = json.dumps(trigger)
                        triggers_list.append(str(trigger))
                        trigger_to_json = json.loads(trigger_to_json)
                        print (str(trigger_to_json["description"]))
                        worksheet["E"+str(cell_count)] = str(trigger_to_json["description"])
                        worksheet["E"+str(cell_count)].font = Font(size=9)
                        worksheet["H"+str(cell_count)] = str(trigger_to_json["expression"])
                        worksheet["H"+str(cell_count)].font = Font(size=9)
                        worksheet["G"+str(cell_count)] = 'Enable' if str(trigger_to_json["status"]) == '0' else 'Disable'
                        worksheet["G"+str(cell_count)].font = Font(size=9)
                        if trigger_to_json["priority"] == "0":
                            severity = "not classified"
                        elif trigger_to_json["priority"] == "1":
                            severity = "information"
                        elif trigger_to_json["priority"] == "2":
                            severity = "warning"
                        elif trigger_to_json["priority"] == "3":
                            severity = "average"
                        elif trigger_to_json["priority"] == "4":
                            severity = "high"
                        elif trigger_to_json["priority"] == "5":
                            severity = "disaster"
                        worksheet["F"+str(cell_count)] = str(severity)
                        worksheet["F"+str(cell_count)].font = Font(size=9)
#                    print (len(item_triggers_list))
                    trigger_count += 1
                    if len(item_triggers_list) > trigger_count:
                        cell_count += 1              
                audit_file.save(filename = dest_filename)
                cell_count += 1
                print ('----------')              
            else:
                None

    audit_file = load_workbook(filename = dest_filename)
    worksheet = audit_file[hostname]

    for column_cells in worksheet.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            worksheet.column_dimensions[new_column_letter].width = new_column_length
    audit_file.save(filename = dest_filename)
    print ('-----//-----')          
zapi.user.logout()
